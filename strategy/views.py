from django.http import HttpResponse
import json
from smtplib import SMTPResponseException
from django.conf import settings
from django.http import JsonResponse
from django.template import loader
from bson.objectid import ObjectId
import redis
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

import pymongo
myclient = pymongo.MongoClient(
    'mongodb://user:pass@host.docker.internal:27017/?authSource=admin&readPreference=primary&appname=MongoDB%20Compass&directConnection=true&ssl=false')
matches = myclient["statistics"]["matches"]
fixtures = myclient["statistics"]["fixtures"]
relations = myclient["statistics"]["relations"]
strategy = myclient["statistics"]["predicts"]
queue = myclient["statistics"]["queue"]
r = redis.Redis(host='host.docker.internal', port=6379, decode_responses=True)


def build_xml():
    return 'hello.xlsx'


def predicts(request, userId):
    return JsonResponse({'data': [{**x, '_id': str(x['_id'])} for x in list(strategy.find({'userId': userId}))]})


def dowload_xml(request, userId):
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Data.xlsx"'

    white_color = openpyxl.styles.colors.Color(rgb='00FFFFFF')
    black_color = openpyxl.styles.colors.Color(rgb='00000000')
    white_font = Font(color=white_color, bold=True)
    black_font = Font(color=black_color, bold=True)
    header_color = openpyxl.styles.colors.Color(rgb='00339966')
    header_fill = openpyxl.styles.fills.PatternFill(
        patternType='solid', fgColor=header_color)
    header_color_2 = openpyxl.styles.colors.Color(rgb='00969696')
    header_fill_2 = openpyxl.styles.fills.PatternFill(
        patternType='solid', fgColor=header_color_2)
    data_color = openpyxl.styles.colors.Color(rgb='00CCFFCC')
    data_fill = openpyxl.styles.fills.PatternFill(
        patternType='solid', fgColor=data_color)
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    predicts = [{**x, '_id': str(x['_id'])}
                for x in list(strategy.find({'userId': userId}))]
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Absolute'
    filtered_groups = {}
    for i, predict in enumerate(predicts):
        filtered_groups = {**filtered_groups, **predict['odds']}

    for i, predict in enumerate(predicts):
        league_cell = sheet.cell(row=i * 9 + 1, column=1)
        league_cell.value = predict['league']
        # league_cell.fill = data_fill
        league_cell.font = black_font

        match_cell = sheet.cell(row=i * 9 + 2, column=1)
        match_cell.value = predict['homeTeam'] + ' - ' + predict['awayTeam']
        # match_cell.fill = data_fill
        # match_cell.font = black_font

        app_cell = sheet.cell(row=i * 9 + 3, column=1)
        app_cell.value = 'App'
        # app_cell.fill = data_fill
        # app_cell.font = black_font

        book_cell = sheet.cell(row=i * 9 + 4, column=1)
        book_cell.value = 'Bookmaker'
        # book_cell.fill = data_fill
        # book_cell.font = black_font

        x_payoff_cell = sheet.cell(row=i * 9 + 5, column=1)
        x_payoff_cell.value = 'Expected payoff'
        # x_payoff_cell.fill = data_fill
        # x_payoff_cell.font = black_font

        match_result_cell = sheet.cell(row=i * 9 + 6, column=1)
        match_result_cell.value = 'Bet'
        # match_result_cell.fill = data_fill
        # match_result_cell.font = black_font

        income_cell = sheet.cell(row=i * 9 + 7, column=1)
        income_cell.value = 'Match result'

        income_cell = sheet.cell(row=i * 9 + 8, column=1)
        income_cell.value = 'Income'
        # income_cell.fill = data_fill
        # income_cell.font = black_font

        # groups
        start_column = 3
        for group_index, group in enumerate(filtered_groups):
            group_cell = sheet.cell(row=i * 9 + 1, column=start_column)
            group_cell.value = group
            group_cell.font = black_font
            total_value = '= 0'
            for odd_index, oddName in enumerate(filtered_groups[group]['hints']):
                odd_name_cell = sheet.cell(
                    row=i * 9 + 2, column=start_column + odd_index)
                odd_name_cell.value = filtered_groups[group]['hints'][odd_index]
                odd_name_cell.fill = header_fill_2
                odd_name_cell.border = thin_border
                odd_name_cell.font = white_font

                odd_relative_cell = sheet.cell(
                    row=i * 9 + 3, column=start_column + odd_index)
                odd_relative_cell.value = predict['odds'][group]['absolute'][odd_index] if group in predict['odds'] else 0
                if group in predict['odds']:
                    odd_relative_cell.fill = data_fill
                odd_relative_cell.border = thin_border

                odd_book_cell = sheet.cell(
                    row=i * 9 + 4, column=start_column + odd_index)
                odd_book_cell.value = 0
                if group in predict['odds']:
                    odd_book_cell.fill = data_fill
                odd_book_cell.border = thin_border

                x_payoff_cell = sheet.cell(
                    row=i * 9 + 5, column=start_column + odd_index)
                x_payoff_cell.value = f'={odd_book_cell.coordinate}/{odd_relative_cell.coordinate}'
                if group in predict['odds']:
                    x_payoff_cell.fill = data_fill
                x_payoff_cell.border = thin_border

                bet_cell = sheet.cell(
                    row=i * 9 + 6, column=start_column + odd_index)
                bet_cell.value = 0
                if group in predict['odds']:
                    bet_cell.fill = data_fill
                bet_cell.border = thin_border

                match_res_cell = sheet.cell(
                    row=i * 9 + 7, column=start_column + odd_index)
                match_res_cell.value = 0
                if group in predict['odds']:
                    match_res_cell.fill = data_fill
                match_res_cell.border = thin_border

                income_cell = sheet.cell(
                    row=i * 9 + 8, column=start_column + odd_index)
                income_cell.value = f'={odd_book_cell.coordinate}*{match_res_cell.coordinate}*{bet_cell.coordinate}-{bet_cell.coordinate}'
                if group in predict['odds']:
                    income_cell.fill = data_fill
                income_cell.border = thin_border

                total_value = f'{total_value}+{income_cell.coordinate}'

            total_cell = sheet.cell(
                row=i * 9 + 1, column=start_column - 1 + len(filtered_groups[group]['hints']))
            total_cell.font = black_font
            total_cell.value = total_value if i == 0 else f'{total_value} + {sheet.cell(row=i * 9 - 9 + 1, column=start_column - 1 + len(filtered_groups[group]["hints"])).coordinate}'
            start_column += len(filtered_groups[group]['hints']) + 1

    sheet = wb.create_sheet('Relative')

    for i, predict in enumerate(predicts):
        league_cell = sheet.cell(row=i * 9 + 1, column=1)
        league_cell.value = predict['league']
        # league_cell.fill = data_fill
        league_cell.font = black_font

        match_cell = sheet.cell(row=i * 9 + 2, column=1)
        match_cell.value = predict['homeTeam'] + ' - ' + predict['awayTeam']
        # match_cell.fill = data_fill
        # match_cell.font = black_font

        app_cell = sheet.cell(row=i * 9 + 3, column=1)
        app_cell.value = 'App'
        # app_cell.fill = data_fill
        # app_cell.font = black_font

        book_cell = sheet.cell(row=i * 9 + 4, column=1)
        book_cell.value = 'Bookmaker'
        # book_cell.fill = data_fill
        # book_cell.font = black_font

        x_payoff_cell = sheet.cell(row=i * 9 + 5, column=1)
        x_payoff_cell.value = 'Expected payoff'
        # x_payoff_cell.fill = data_fill
        # x_payoff_cell.font = black_font

        match_result_cell = sheet.cell(row=i * 9 + 6, column=1)
        match_result_cell.value = 'Bet'
        # match_result_cell.fill = data_fill
        # match_result_cell.font = black_font

        income_cell = sheet.cell(row=i * 9 + 7, column=1)
        income_cell.value = 'Match result'

        income_cell = sheet.cell(row=i * 9 + 8, column=1)
        income_cell.value = 'Income'
        # income_cell.fill = data_fill
        # income_cell.font = black_font

        # groups
        start_column = 3
        for group_index, group in enumerate(filtered_groups):
            group_cell = sheet.cell(row=i * 9 + 1, column=start_column)
            group_cell.value = group
            group_cell.font = black_font
            total_value = '= 0'
            for odd_index, oddName in enumerate(filtered_groups[group]['hints']):
                odd_name_cell = sheet.cell(
                    row=i * 9 + 2, column=start_column + odd_index)
                odd_name_cell.value = filtered_groups[group]['hints'][odd_index]
                odd_name_cell.fill = header_fill_2
                odd_name_cell.border = thin_border
                odd_name_cell.font = white_font

                odd_relative_cell = sheet.cell(
                    row=i * 9 + 3, column=start_column + odd_index)
                odd_relative_cell.value = predict['odds'][group]['relative'][odd_index] if group in predict['odds'] else 0
                if group in predict['odds']:
                    odd_relative_cell.fill = data_fill
                odd_relative_cell.border = thin_border

                odd_book_cell = sheet.cell(
                    row=i * 9 + 4, column=start_column + odd_index)
                odd_book_cell.value = 0
                if group in predict['odds']:
                    odd_book_cell.fill = data_fill
                odd_book_cell.border = thin_border

                x_payoff_cell = sheet.cell(
                    row=i * 9 + 5, column=start_column + odd_index)
                x_payoff_cell.value = f'={odd_book_cell.coordinate}/{odd_relative_cell.coordinate}'
                if group in predict['odds']:
                    x_payoff_cell.fill = data_fill
                x_payoff_cell.border = thin_border

                bet_cell = sheet.cell(
                    row=i * 9 + 6, column=start_column + odd_index)
                bet_cell.value = 0
                if group in predict['odds']:
                    bet_cell.fill = data_fill
                bet_cell.border = thin_border

                match_res_cell = sheet.cell(
                    row=i * 9 + 7, column=start_column + odd_index)
                match_res_cell.value = 0
                if group in predict['odds']:
                    match_res_cell.fill = data_fill
                match_res_cell.border = thin_border

                income_cell = sheet.cell(
                    row=i * 9 + 8, column=start_column + odd_index)
                income_cell.value = f'={odd_book_cell.coordinate}*{match_res_cell.coordinate}*{bet_cell.coordinate}-{bet_cell.coordinate}'
                if group in predict['odds']:
                    income_cell.fill = data_fill
                income_cell.border = thin_border

                total_value = f'{total_value}+{income_cell.coordinate}'

            total_cell = sheet.cell(
                row=i * 9 + 1, column=start_column + len(filtered_groups[group]['hints']))
            total_cell.value = total_value if i == 0 else f'{total_value} + {sheet.cell(row=i * 9 - 9 + 1, column=start_column - 1 + len(filtered_groups[group]["hints"])).coordinate}'
            start_column += len(filtered_groups[group]['hints']) + 1

    wb.save(response)

    return response


def add_to_queue(request, userId):
    body_unicode = request.body.decode('utf-8')
    body = json.loads(body_unicode)
    body['userId'] = int(body['userId'])
    # res = queue.update_one({'user_id': body['userId']}, {
    #    '$push': {'queue': [body['homeTeam'], body['awayTeam']]}}, upsert=True)
    res = queue.insert_one(body)
    print(res.inserted_id)
    r.publish('task', json.dumps({'task_id': str(res.inserted_id)}))
    return JsonResponse({'data': json.loads(json.dumps(list(queue.find({'userId': int(userId)})), default=str))})


def get_queue(request, userId):
    print(list(queue.find({'userId': userId}, projection={'_id': 0})))
    return JsonResponse({'data': json.loads(json.dumps(list(queue.find({'userId': userId})), default=str))})


def delete_from_queue(request, matchId):
    body_unicode = request.body.decode('utf-8')
    body = json.loads(body_unicode)
    match = queue.delete_one({'_id': ObjectId(matchId)})
    return JsonResponse({'data': json.loads(json.dumps(list(queue.find({'userId': int(body['userId'])})), default=str))})
    # return JsonResponse({'data': [x for x in list(strategy.find(projection={'_id': 0}))]})


def delete_from_results(request, matchId):
    body_unicode = request.body.decode('utf-8')
    body = json.loads(body_unicode)
    match = strategy.delete_one({'_id': ObjectId(matchId)})
    return JsonResponse({'data': [{**x, '_id': str(x['_id'])} for x in list(strategy.find({'userId': int(body['userId'])}))]})
