
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font


def generate_xl_file(strategy, user_id, selected_leagues, selected_groups):

    white_color = openpyxl.styles.colors.Color(rgb='00FFFFFF')
    black_color = openpyxl.styles.colors.Color(rgb='00000000')
    white_font = Font(color=white_color, bold=True)
    black_font = Font(color=black_color, bold=True)
    header_color = openpyxl.styles.colors.Color(rgb='00331010')
    header_fill = openpyxl.styles.fills.PatternFill(
        patternType='solid', fgColor=header_color)
    header_color_2 = openpyxl.styles.colors.Color(rgb='00106106')
    header_fill_2 = openpyxl.styles.fills.PatternFill(
        patternType='solid', fgColor=header_color_2)
    data_color = openpyxl.styles.colors.Color(rgb='00CCFFCC')
    data_fill = openpyxl.styles.fills.PatternFill(
        patternType='solid', fgColor=data_color)
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    query = {
        "userId": user_id,
        "league": {"$in": selected_leagues},
        "$or": [{f"odds.{k}": {"$exists": True}} for k in selected_groups]
    }

    predicts = [{**x, '_id': str(x['_id'])}
                for x in list(strategy.find(query))]
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Absolute'
    filtered_groups = {}
    for i, predict in enumerate(predicts):
        filtered_groups = {**filtered_groups, **{k: v for k, v in predict['odds'].items() if k in selected_groups}}

    for i, predict in enumerate(predicts):
        predict['odds'] = {k: v for k, v in predict['odds'].items() if k in selected_groups}
        league_cell = sheet.cell(row=i * 10 + 1, column=1)
        league_cell.value = predict['league']
        # league_cell.fill = data_fill
        league_cell.font = black_font

        match_cell = sheet.cell(row=i * 10 + 2, column=1)
        match_cell.value = predict['homeTeam'] + ' - ' + predict['awayTeam']
        # match_cell.fill = data_fill
        # match_cell.font = black_font

        app_cell = sheet.cell(row=i * 10 + 3, column=1)
        app_cell.value = 'App'
        # app_cell.fill = data_fill
        # app_cell.font = black_font

        book_cell = sheet.cell(row=i * 10 + 4, column=1)
        book_cell.value = 'Bookmaker'
        # book_cell.fill = data_fill
        # book_cell.font = black_font

        x_payoff_cell = sheet.cell(row=i * 10 + 5, column=1)
        x_payoff_cell.value = 'Expected payoff'
        # x_payoff_cell.fill = data_fill
        # x_payoff_cell.font = black_font

        match_result_cell = sheet.cell(row=i * 10 + 6, column=1)
        match_result_cell.value = 'Bet'
        # match_result_cell.fill = data_fill
        # match_result_cell.font = black_font

        income_cell = sheet.cell(row=i * 10 + 7, column=1)
        income_cell.value = 'Match result'

        income_cell = sheet.cell(row=i * 10 + 8, column=1)
        income_cell.value = 'Income'
        # income_cell.fill = data_fill
        # income_cell.font = black_font

        income_cell = sheet.cell(row=i * 9 + 9, column=1)
        income_cell.value = 'Test'

        # groups
        start_column = 3
        for group_index, group in enumerate(filtered_groups):
            group_cell = sheet.cell(row=i * 10 + 1, column=start_column)
            group_cell.value = group
            group_cell.font = black_font
            total_value = '= 0'
            for odd_index, oddName in enumerate(filtered_groups[group]['hints']):
                odd_name_cell = sheet.cell(
                    row=i * 10 + 2, column=start_column + odd_index)
                odd_name_cell.value = filtered_groups[group]['hints'][odd_index]
                odd_name_cell.fill = header_fill_2
                odd_name_cell.border = thin_border
                odd_name_cell.font = white_font

                odd_relative_cell = sheet.cell(
                    row=i * 10 + 3, column=start_column + odd_index)
                odd_relative_cell.value = predict['odds'][group]['absolute'][odd_index] if group in predict['odds'] else 0
                if group in predict['odds']:
                    odd_relative_cell.fill = data_fill
                odd_relative_cell.border = thin_border

                odd_book_cell = sheet.cell(
                    row=i * 10 + 4, column=start_column + odd_index)
                odd_book_cell.value = 0
                if group in predict['odds']:
                    odd_book_cell.fill = data_fill
                odd_book_cell.border = thin_border

                x_payoff_cell = sheet.cell(
                    row=i * 10 + 5, column=start_column + odd_index)
                x_payoff_cell.value = f'={odd_book_cell.coordinate}/{odd_relative_cell.coordinate}'
                if group in predict['odds']:
                    x_payoff_cell.fill = data_fill
                x_payoff_cell.border = thin_border

                bet_cell = sheet.cell(
                    row=i * 10 + 6, column=start_column + odd_index)
                bet_cell.value = 0
                if group in predict['odds']:
                    bet_cell.fill = data_fill
                bet_cell.border = thin_border

                match_res_cell = sheet.cell(
                    row=i * 10 + 7, column=start_column + odd_index)
                match_res_cell.value = 0
                if group in predict['odds']:
                    match_res_cell.fill = data_fill
                match_res_cell.border = thin_border

                income_cell = sheet.cell(
                    row=i * 10 + 8, column=start_column + odd_index)
                income_cell.value = f'={odd_book_cell.coordinate}*{match_res_cell.coordinate}*{bet_cell.coordinate}-{bet_cell.coordinate}'
                if group in predict['odds']:
                    income_cell.fill = data_fill
                income_cell.border = thin_border

                test_cell = sheet.cell(
                    row=i * 10 + 9, column=start_column + odd_index)
                test_cell.value = f'={odd_relative_cell.coordinate}*{match_res_cell.coordinate}*{bet_cell.coordinate}-{bet_cell.coordinate}'
                if group in predict['odds']:
                    test_cell.fill = data_fill
                test_cell.border = thin_border

                total_value = f'{total_value}+{income_cell.coordinate}'

            total_cell = sheet.cell(
                row=i * 10 + 1, column=start_column - 1 + len(filtered_groups[group]['hints']))
            total_cell.font = black_font
            total_cell.value = total_value if i == 0 else f'{total_value} + {sheet.cell(row=i * 10 - 10 + 1, column=start_column - 1 + len(filtered_groups[group]["hints"])).coordinate}'
            start_column += len(filtered_groups[group]['hints']) + 1

    sheet = wb.create_sheet('Relative')

    for i, predict in enumerate(predicts):
        predict['odds'] = {k: v for k, v in predict['odds'].items() if k in selected_groups}
        league_cell = sheet.cell(row=i * 10 + 1, column=1)
        league_cell.value = predict['league']
        # league_cell.fill = data_fill
        league_cell.font = black_font

        match_cell = sheet.cell(row=i * 10 + 2, column=1)
        match_cell.value = predict['homeTeam'] + ' - ' + predict['awayTeam']
        # match_cell.fill = data_fill
        # match_cell.font = black_font

        app_cell = sheet.cell(row=i * 10 + 3, column=1)
        app_cell.value = 'App'
        # app_cell.fill = data_fill
        # app_cell.font = black_font

        book_cell = sheet.cell(row=i * 10 + 4, column=1)
        book_cell.value = 'Bookmaker'
        # book_cell.fill = data_fill
        # book_cell.font = black_font

        x_payoff_cell = sheet.cell(row=i * 10 + 5, column=1)
        x_payoff_cell.value = 'Expected payoff'
        # x_payoff_cell.fill = data_fill
        # x_payoff_cell.font = black_font

        match_result_cell = sheet.cell(row=i * 10 + 6, column=1)
        match_result_cell.value = 'Bet'
        # match_result_cell.fill = data_fill
        # match_result_cell.font = black_font

        income_cell = sheet.cell(row=i * 10 + 7, column=1)
        income_cell.value = 'Match result'

        income_cell = sheet.cell(row=i * 10 + 8, column=1)
        income_cell.value = 'Income'
        # income_cell.fill = data_fill
        # income_cell.font = black_font

        income_cell = sheet.cell(row=i * 10 + 9, column=1)
        income_cell.value = 'Test'

        # groups
        start_column = 3
        for group_index, group in enumerate(filtered_groups):

            #League
            group_cell = sheet.cell(row=i * 10 + 1, column=start_column)
            group_cell.value = group
            group_cell.font = black_font
            total_value = '= 0'
            test_value = '= 0'
            relative_bets = '(0'
            bet_cells = []
            xp_cells = []

            for odd_index, oddName in enumerate(filtered_groups[group]['hints']):

                #Teams
                odd_name_cell = sheet.cell(
                    row=i * 10 + 2, column=start_column + odd_index)
                odd_name_cell.value = filtered_groups[group]['hints'][odd_index]
                odd_name_cell.fill = header_fill_2
                odd_name_cell.border = thin_border
                odd_name_cell.font = white_font

                #App
                odd_relative_cell = sheet.cell(
                    row=i * 10 + 3, column=start_column + odd_index)
                odd_relative_cell.value = predict['odds'][group]['relative'][odd_index] if group in predict['odds'] else 0
                if group in predict['odds']:
                    odd_relative_cell.fill = data_fill
                odd_relative_cell.border = thin_border

                #Book
                odd_book_cell = sheet.cell(
                    row=i * 10 + 4, column=start_column + odd_index)
                odd_book_cell.value = 0
                if group in predict['odds']:
                    odd_book_cell.fill = data_fill
                odd_book_cell.border = thin_border

                #X payoff
                x_payoff_cell = sheet.cell(
                    row=i * 10 + 5, column=start_column + odd_index)
                x_payoff_cell.value = f'={odd_book_cell.coordinate}/{odd_relative_cell.coordinate}'
                if group in predict['odds']:
                    x_payoff_cell.fill = data_fill
                x_payoff_cell.border = thin_border

                #Bet
                bet_cell = sheet.cell(
                    row=i * 10 + 6, column=start_column + odd_index)
                relative_bets += f'+IF({x_payoff_cell.coordinate}>=1,({x_payoff_cell.coordinate}),0)'
                bet_cell.value = f'=IF({x_payoff_cell.coordinate}>=1,({x_payoff_cell.coordinate})/(RELATIVE_BETS),0)'
                bet_cells.append(bet_cell)
                if group in predict['odds']:
                    bet_cell.fill = data_fill
                bet_cell.border = thin_border

                #Match result
                match_res_cell = sheet.cell(
                    row=i * 10 + 7, column=start_column + odd_index)
                match_res_cell.value = 0
                if group in predict['odds']:
                    match_res_cell.fill = data_fill
                match_res_cell.border = thin_border

                #Income
                income_cell = sheet.cell(
                    row=i * 10 + 8, column=start_column + odd_index)
                income_cell.value = f'={odd_book_cell.coordinate}*{match_res_cell.coordinate}*{bet_cell.coordinate}-{bet_cell.coordinate}'
                if group in predict['odds']:
                    income_cell.fill = data_fill
                income_cell.border = thin_border

                #Test
                test_cell = sheet.cell(
                    row=i * 10 + 9, column=start_column + odd_index)
                test_cell.value = f'={odd_relative_cell.coordinate}*{match_res_cell.coordinate}*{bet_cell.coordinate}-{bet_cell.coordinate}'
                if group in predict['odds']:
                    test_cell.fill = data_fill
                test_cell.border = thin_border

                total_value = f'{total_value}+{income_cell.coordinate}'
                test_value = f'{test_value}+{test_cell.coordinate}'

            #Total income
            total_cell = sheet.cell(
                row=i * 10 + 1, column=start_column - 1 + len(filtered_groups[group]['hints']))
            total_cell.font = black_font
            total_cell.value = total_value if i == 0 else f'{total_value} + {sheet.cell(row=i * 10 - 10 + 1, column=start_column - 1 + len(filtered_groups[group]["hints"])).coordinate}'

            #Test income
            sheet.cell(
                row=i * 10 + 10, column=start_column).value = 'Luck coefficient'
            luck_cell = sheet.cell(
                row=i * 10 + 10, column=start_column - 1 + len(filtered_groups[group]['hints']))
            #luck_cell.font = black_font
            luck_cell.value = test_value if i == 0 else f'{test_value} + {sheet.cell(row=i * 10, column=start_column - 1 + len(filtered_groups[group]["hints"])).coordinate}'
            
            start_column += len(filtered_groups[group]['hints']) + 1
            
            relative_bets += ')'
            for bet_cell in bet_cells:
                bet_cell.value = bet_cell.value.replace('RELATIVE_BETS', relative_bets)
                #print(bet_cell.value)
    return wb
